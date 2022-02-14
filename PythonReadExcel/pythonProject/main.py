import openpyxl




wb_obj = openpyxl.load_workbook('Podaci.xlsx')

df = wb_obj.worksheets[1]

f= open("Users.txt","w+", encoding='utf8')
#new User { IDUser = "S00001", Owner = "Novakovic Petar", Password = "317D" },
for row in range(df.max_row):
       # print('\n')
        f.write('\n')
        f.write("new User { IDUser = \"")
        #f.write("new User { IDUser = "S00001", Owner = "Novakovic Petar", Password = "317D" },")
        for column in range(df.max_column):
            #print(df.cell(row+1,column+1).value)
            f.write(str(df.cell(row+1,column+1).value))
            if column == 0:
                f.write("\", Owner = \"")
            elif column ==1:
                f.write("\", Password = \"")
            else:
                f.write("\" },")



f.close()


wb_obj = openpyxl.load_workbook('Podaci.xlsx')

df = wb_obj.worksheets[0]

f= open("Traffics.txt","w+", encoding='utf8')

cnt = 5
#new Traffic { ID = 1, UserId = "S00001", Date = new DateTime(2021,1,1), Document= "Početno stanje", Owes= 580, Claims = 0 },
for row in range(df.max_row):
       # print('\n')
        f.write('\n')
        f.write("new Traffic { ID = "+str(cnt)+", UserId = \"")
        cnt+=1
        for column in range(df.max_column):
            #print(df.cell(row+1,column+1).value)
            if column !=1:
                f.write(str(df.cell(row+1,column+1).value))
            else:
                pre_parse = str(df.cell(row+1,column+1).value).split(' ')
                pre_parse = pre_parse[0].split('-')
                if len(pre_parse) > 2:
                    f.write("new DateTime("+pre_parse[0]+", "+pre_parse[1]+", "+pre_parse[2]+")")

            if column == 0:
                f.write("\", Date = ")
            elif column ==1:
                f.write(", Document = \"")
            elif column ==2:
                f.write("\", Owes = ")
            elif column ==3:
                f.write(", Claims = ")
            else:
                f.write(" },")



f.close()